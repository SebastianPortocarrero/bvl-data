#!/usr/bin/env python3
"""
procesar_aporte.py — INCORPORA UN APORTE A LA BASE, SIN DEPENDER DE NINGUNA MAC.

Este script hace autosuficiente al repo bvl-data: toma un .zip (o una carpeta) con
Excel oficiales de SMV — los que genera el script descargar_smv.py — y actualiza
{TICKER}.json + index.json de este mismo repo.

Pensado para correr en un entorno de nube (routine de claude.ai) o en cualquier
máquina con Python + openpyxl. NO necesita Selenium ni acceso a smv.gob.pe: el
aporte ya trae los trimestres descargados.

Uso:
    python3 procesar_aporte.py ARCHIVO.zip --ticker SNJUANC1
    python3 procesar_aporte.py carpeta/ --ticker SNJUANC1 --tipo Consolidada

Si se omite --ticker, se intenta resolver por el nombre de la empresa contra
TradingView (prefiere la clase común C1).

El parser de abajo es copia literal del validado en el sistema local: respeta las
trampas de los EEFF de SMV (deuda financiera en DOS líneas, utilidad neta duplicada
en el flujo de efectivo, cifras atribuibles a la controladora, miles→millones,
etiquetas distintas en bancos).
"""
import argparse
import datetime
import glob
import json
import os
import re
import sys
import tempfile
import urllib.parse
import urllib.request
import zipfile

import openpyxl

QUARTERS = ["TRIMESTRE I", "TRIMESTRE II", "TRIMESTRE III", "TRIMESTRE IV"]
REPO_DIR = os.path.dirname(os.path.abspath(__file__))


def read_quarter_excel(xlsx_path):
    """Lee un Excel de SMV y retorna dict con las líneas clave del Estado de Resultados y Balance.

    SMV reporta todos los montos en MILES de soles — se dividen entre 1000 aquí para trabajar
    siempre en MILLONES de soles de forma consistente en toda la skill.

    Algunas etiquetas (ej. "Otros Pasivos Financieros") aparecen DOS VECES: una en Pasivos
    Corrientes y otra en Pasivos No Corrientes. Un dict simple {label: row} se queda solo con
    la última ocurrencia, por eso aquí se recorre en orden detectando la sección de cada una.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["ReporteEstadosFinancieros"]
    all_rows = [r for r in ws.iter_rows(values_only=True) if r[0]]
    rows = {}
    otros_pasivos_fin = []
    section = None
    # "Propietarios de la Controladora" aparece en varios estados; la que vale es la
    # PRIMERA inmediatamente después de "Ganancia (Pérdida) Neta del Ejercicio"
    # (desglose de la utilidad entre controladora y no controladoras).
    utilidad_ctrl_row = None
    tras_ganancia_neta = False
    # "Ganancia (Pérdida) Neta del Ejercicio" reaparece en el Estado de Flujo de
    # Efectivo (siempre YTD acumulado) DESPUÉS del Estado de Resultados. Como
    # rows[label]=r se queda con la última ocurrencia, val() terminaría leyendo el
    # acumulado del flujo de caja en vez del trimestre discreto. Guardamos la
    # PRIMERA ocurrencia (la del Estado de Resultados, columna 2 = período de 3 meses).
    utilidad_neta_row = None
    for r in all_rows:
        label = str(r[0]).strip()
        rows[label] = r
        if label == "Pasivos Corrientes":
            section = "corriente"
        elif label == "Pasivos No Corrientes":
            section = "no_corriente"
        if label == "Otros Pasivos Financieros":
            otros_pasivos_fin.append((section, r))
        if label == "Ganancia (Pérdida) Neta del Ejercicio":
            if utilidad_neta_row is None:
                utilidad_neta_row = r
            tras_ganancia_neta = True
        elif tras_ganancia_neta and label == "Propietarios de la Controladora":
            if utilidad_ctrl_row is None:
                utilidad_ctrl_row = r
            tras_ganancia_neta = False
        elif tras_ganancia_neta and "atribuible a" in label.lower():
            pass  # fila intermedia "Ganancia (Pérdida) Neta atribuible a:" — seguir
        elif tras_ganancia_neta:
            tras_ganancia_neta = False

    def parse_num(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", "").replace("(", "-").replace(")", ""))
        except Exception:
            return 0.0

    MILES_A_MILLONES = 1000.0

    def val(label, col=2, scale=True):
        row = rows.get(label)
        if row is None:
            return 0.0
        v = row[col] if len(row) > col else None
        n = parse_num(v)
        return n / MILES_A_MILLONES if scale else n

    def val_any(labels, col=2, scale=True):
        """Primer alias presente con valor ≠ 0; si todos son 0, el primero presente.

        Los bancos/financieras (ej. CREDITC1) no usan las etiquetas industriales:
        ingresos = 'Ingreso por Intereses', y no existe 'TOTAL ACTIVOS' (se usa
        'TOTAL PASIVO Y PATRIMONIO', idéntico por identidad contable).
        """
        first_present = 0.0
        found = False
        for label in labels:
            if label in rows:
                n = val(label, col, scale)
                if not found:
                    first_present, found = n, True
                if n != 0.0:
                    return n
        return first_present

    eps_q = val("Total de Ganancias (Pérdida) Básica por Acción Ordinaria", scale=False)

    # D&A: solo aparece si la empresa reporta flujo por método indirecto (la mayoría
    # en SMV usa método directo → no hay línea y EBITDA queda como N/D).
    dya_q = 0.0
    for label, r in rows.items():
        if "depreciaci" in label.lower() and "amortizaci" in label.lower():
            dya_q = abs(parse_num(r[2] if len(r) > 2 else None)) / MILES_A_MILLONES
            break

    deuda_fin_corriente = 0.0
    deuda_fin_no_corriente = 0.0
    for sec, r in otros_pasivos_fin:
        n = parse_num(r[2] if len(r) > 2 else None) / MILES_A_MILLONES
        if sec == "corriente":
            deuda_fin_corriente += n
        elif sec == "no_corriente":
            deuda_fin_no_corriente += n

    es_financiera = ("Ingreso por Intereses" in rows and
                     val("Ingresos de Actividades Ordinarias") == 0.0)

    # Utilidad neta del trimestre (3 meses) desde la PRIMERA fila del Estado de
    # Resultados, evitando el YTD de la fila homónima del Estado de Flujo de Efectivo.
    _utilidad_total_q = (parse_num(utilidad_neta_row[2]) / MILES_A_MILLONES
                         if utilidad_neta_row is not None and len(utilidad_neta_row) > 2
                         else 0.0)

    return {
        "activos_totales": val_any(["TOTAL ACTIVOS", "TOTAL PASIVO Y PATRIMONIO"]),
        "activo_corriente": val("Total Activos Corrientes"),
        "total_pasivos": val("Total Pasivos"),
        "pasivo_corriente": val("Total Pasivos Corrientes"),
        "patrimonio": val("Total Patrimonio"),
        "caja": val("Efectivo y Equivalentes al Efectivo"),
        "inventarios": val("Inventarios"),
        "eps_q": eps_q,
        "es_financiera": es_financiera,

        "ingresos_q": val_any(["Ingresos de Actividades Ordinarias", "Ingreso por Intereses"]),
        "costo_ventas_q": abs(val_any(["Costo de Ventas", "Gasto por Intereses"])),
        "gastos_ventas_q": abs(val("Gastos de Ventas y Distribución")),
        "gastos_admin_q": abs(val("Gastos de Administración")),
        "otros_ing_op_q": val("Otros Ingresos Operativos"),
        "otros_gas_op_q": abs(val("Otros Gastos Operativos")),
        "ing_financieros_q": val("Ingresos Financieros"),
        "gas_financieros_q": abs(val("Gastos Financieros")),
        "utilidad_q": _utilidad_total_q,
        # Atribuible a la controladora (excluye minoritarios) — es la base correcta
        # para UPA/VPA/ROE y lo que usan TradingView/Investidor10. Si la empresa no
        # desglosa controladora (sin minoritarios, ej. Cerro Verde), el fallback es la
        # utilidad total del Estado de Resultados (3 meses), NUNCA el val() acumulado
        # que quedaría contaminado por la fila homónima del flujo de caja (YTD).
        "utilidad_ctrl_q": (parse_num(utilidad_ctrl_row[2]) / MILES_A_MILLONES
                            if utilidad_ctrl_row is not None and len(utilidad_ctrl_row) > 2
                            else _utilidad_total_q),
        "patrimonio_ctrl": (val("Patrimonio Atribuible a los Propietarios de la Controladora")
                            or val("Total Patrimonio")),
        "dya_q": dya_q,

        "deuda_fin_corriente": deuda_fin_corriente,
        "deuda_fin_no_corriente": deuda_fin_no_corriente,
    }


def ebit_of_quarter(d):
    """EBIT aproximado de un trimestre a partir de sus líneas del Estado de Resultados."""
    return (d.get("ingresos_q", 0) - d.get("costo_ventas_q", 0)
            - d.get("gastos_ventas_q", 0) - d.get("gastos_admin_q", 0)
            + d.get("otros_ing_op_q", 0) - d.get("otros_gas_op_q", 0))


def save_ttm_to_excel(ticker, anio_ref, trimestre_ref, tipo, ttm, outdir):
    """Guarda los datos TTM calculados en la hoja del ticker en indicadores.xlsx."""
    os.makedirs(outdir, exist_ok=True)
    excel = os.path.join(outdir, "indicadores.xlsx")
    try:
        wb = openpyxl.load_workbook(excel)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    from openpyxl.styles import Font, PatternFill
    ttm_headers = ["fecha_corrida", "anio_smv", "trimestre_smv", "tipo_eeff",
                   "ingresos_ttm_M", "utilidad_ttm_M", "ebit_aprox_ttm_M",
                   "activos_totales_M", "activo_corriente_M", "patrimonio_M",
                   "deuda_fin_M", "deuda_neta_M", "caja_M",
                   "eps_ttm", "vpa"]

    if ticker in wb.sheetnames:
        existing_headers = [c.value for c in wb[ticker][1]]
        if "anio_smv" not in existing_headers:
            # Hoja de un esquema viejo/manual (previo al script TTM) — se conserva como respaldo
            # y se crea una hoja nueva limpia con el esquema TTM correcto.
            backup_name = f"{ticker}_legacy"
            i = 1
            while backup_name in wb.sheetnames:
                backup_name = f"{ticker}_legacy{i}"
                i += 1
            wb[ticker].title = backup_name
            print(f"  ℹ Hoja '{ticker}' tenía un esquema antiguo; renombrada a '{backup_name}' y se crea una nueva.")

    if ticker not in wb.sheetnames:
        ws = wb.create_sheet(ticker)
        ws.append(ttm_headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="003F87")
            c.fill = PatternFill("solid", fgColor="DCE7F7")
    else:
        ws = wb[ticker]

    # No duplicar misma combinación anio+trimestre
    headers = [c.value for c in ws[1]]
    ai, ti = headers.index("anio_smv"), headers.index("trimestre_smv")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ai] == anio_ref and row[ti] == trimestre_ref:
            print(f"  ℹ Ya existe fila {trimestre_ref} {anio_ref} en Excel, se omite duplicar.")
            wb.save(excel)
            return

    ws.append([
        datetime.date.today().isoformat(), anio_ref, trimestre_ref, tipo,
        round(ttm["ingresos_ttm"], 2), round(ttm["utilidad_ttm"], 2), round(ttm["ebit_ttm"], 2),
        round(ttm["activos_totales"], 2), round(ttm["activo_corriente"], 2), round(ttm["patrimonio"], 2),
        round(ttm["deuda_fin"], 2), round(ttm["deuda_neta"], 2), round(ttm["caja"], 2),
        round(ttm["eps_ttm"], 4), round(ttm["vpa_approx"], 4),
    ])
    wb.save(excel)
    print(f"  ✓ Guardado TTM en indicadores.xlsx, hoja {ticker}")

def ticker_por_nombre(razon):
    q = urllib.parse.quote(" ".join(razon.split()[:3]))
    url = f"https://symbol-search.tradingview.com/symbol_search/?text={q}&exchange=BVL&type=stock"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0",
                                               "Origin": "https://www.tradingview.com"})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            data = json.loads(r.read().decode())
    except Exception:
        return None
    simbolos = [d["symbol"] for d in data if d.get("exchange") == "BVL"]
    for s in simbolos:
        if s.endswith("C1"):
            return s
    return simbolos[0] if simbolos else None


def razon_social_de(xlsx):
    try:
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb["ReporteEstadosFinancieros"]
        for row in ws.iter_rows(max_row=12, values_only=True):
            if row and row[0] and str(row[0]).strip().startswith("Empresa:"):
                return str(row[0]).split("Empresa:", 1)[1].strip()
    except Exception:
        pass
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("entrada", help="archivo .zip o carpeta con los SMV_*.xlsx")
    ap.add_argument("--ticker", default=None)
    ap.add_argument("--tipo", default="Consolidada")
    ap.add_argument("--repo", default=REPO_DIR, help="carpeta del repo bvl-data")
    args = ap.parse_args()

    # 1) reunir los xlsx (de un zip o de una carpeta)
    if args.entrada.lower().endswith(".zip"):
        tmp = tempfile.mkdtemp(prefix="aporte_")
        with zipfile.ZipFile(args.entrada) as zf:
            archivos = [zf.extract(m, tmp) for m in zf.namelist() if m.lower().endswith(".xlsx")]
    else:
        archivos = sorted(glob.glob(os.path.join(args.entrada, "*.xlsx")))
    if not archivos:
        sys.exit("No se encontraron archivos .xlsx en la entrada.")

    # 2) parsear cada trimestre
    empresa, trimestres = None, {}
    for x in archivos:
        m = re.search(r"_(\d{4})_TRIMESTRE(I{1,3}|IV)\.xlsx$", os.path.basename(x))
        if not m:
            print(f"  ⚠ nombre sin periodo, ignorado: {os.path.basename(x)}")
            continue
        empresa = empresa or razon_social_de(x)
        anio, tri = int(m.group(1)), f"TRIMESTRE {m.group(2)}"
        d = read_quarter_excel(x)
        d["ebit_q"] = round(ebit_of_quarter(d), 3)
        trimestres[(anio, tri)] = {
            "anio": anio, "trimestre": tri,
            "orden": anio * 4 + QUARTERS.index(tri),
            "label": f"{QUARTERS.index(tri)+1}T{anio}",
            "datos_M": {k: (round(v, 3) if isinstance(v, float) else v) for k, v in d.items()},
        }
        print(f"  + {m.group(2):>3} {anio}")

    if not trimestres:
        sys.exit("Ningún trimestre válido.")
    if not empresa:
        sys.exit("No pude leer la razón social del encabezado 'Empresa:' de los Excel.")

    ticker = args.ticker or ticker_por_nombre(empresa)
    if not ticker:
        sys.exit(f"No pude resolver el ticker de '{empresa}'. Pasa --ticker XXXXX")

    # 3) fusionar con lo que ya existe en el repo (nunca perder trimestres previos)
    destino = os.path.join(args.repo, f"{ticker}.json")
    if os.path.exists(destino):
        previo = json.load(open(destino, encoding="utf-8"))
        for t in previo.get("trimestres", []):
            clave = (t["anio"], t["trimestre"])
            trimestres.setdefault(clave, t)
        print(f"  ℹ fusionado con {previo.get('n_trimestres', 0)} trimestres ya existentes")

    lista = sorted(trimestres.values(), key=lambda t: t["orden"], reverse=True)
    data = {
        "ticker": ticker, "empresa": empresa, "tipo_eeff": args.tipo,
        "moneda_eeff": "PEN",
        "unidad": "millones (los EEFF de SMV vienen en miles; aquí ya convertidos)",
        "fuente": "SMV — Superintendencia del Mercado de Valores (EEFF oficiales)",
        "actualizado": datetime.date.today().isoformat(),
        "n_trimestres": len(lista), "trimestres": lista,
    }
    with open(destino, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"  ✓ {ticker}.json → {len(lista)} trimestres ({lista[-1]['label']} → {lista[0]['label']})")

    # 4) regenerar index.json leyendo todos los {TICKER}.json del repo
    entradas = []
    for p in sorted(glob.glob(os.path.join(args.repo, "*.json"))):
        if os.path.basename(p) == "index.json":
            continue
        d = json.load(open(p, encoding="utf-8"))
        if "ticker" not in d:
            continue
        ts = d.get("trimestres", [])
        entradas.append({"ticker": d["ticker"], "empresa": d.get("empresa"),
                         "n_trimestres": d.get("n_trimestres", len(ts)),
                         "desde": ts[-1]["label"] if ts else None,
                         "hasta": ts[0]["label"] if ts else None})
    with open(os.path.join(args.repo, "index.json"), "w", encoding="utf-8") as f:
        json.dump({"actualizado": datetime.date.today().isoformat(), "tickers": entradas},
                  f, ensure_ascii=False, indent=1)
    print(f"  ✓ index.json ({len(entradas)} tickers)")
    if len(lista) < 20:
        print(f"  ⚠ {ticker} tiene {len(lista)}/20 trimestres — por debajo del mínimo de 5 años")


if __name__ == "__main__":
    main()
